#!/usr/bin/env python3
"""
xlsx_kpi_engine.py
Smart workbook builder for manufacturing KPI dashboards.
Uses openpyxl for advanced formatting: tables, conditional formatting, named styles.

Usage:
    from xlsx_kpi_engine import generate_kpi_workbook

    result = generate_kpi_dashboard(
        data={
            "Metrics": [["Efficiency %", "85%", "90% Target"], ...],
            "Costs": [["Material Cost", 12500, 15000], ...]
        },
        output_path="/opt/data/workspace/OAKAI_KPI_Dashboard.xlsx",
        title="OAKAI Manufacturing KPI Dashboard",
        subtitle="Generated: 2026-08-12"
    )
"""
import os, sys
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timezone

# ── Brand Styles ──
TEAL_DARK = "0B3D3D"
TEAL      = "0F5C56"
GOLD      = "C69B4B"
CHARCOAL  = "2B2B2B"
GREY      = "6E6E6E"
ROW_ALT   = "F7FAF9"

styles = {}
_stylesheets = ["OAKAI", "OAKAI-Bold"]

def _make_styles(wb):
    s = {}
    s["header"] = NamedStyle(name="header", font=Font(bold=True, color="FFFFFF", size=11),
        fill=PatternFill("solid", fgColor=TEAL_DARK), alignment=Alignment(horizontal="center", vertical="center"),
        border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick')))
    s["value"] = NamedStyle(name="value",
        border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        font=Font(size=10, color=CHARCOAL))
    s["value"].number_format = '#,##0.00'

    s["value_center"] = NamedStyle(name="value_center",
        alignment=Alignment(horizontal="center", vertical="center"),
        border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        font=Font(size=10, color=CHARCOAL))
    s["value_center"].number_format = '#,##0.00'

    s["kpi_good"] = NamedStyle(name="kpi_good", font=Font(color="006100", bold=True),
        fill=PatternFill("solid", fgColor="C6EFCE"))

    s["kpi_bad"] = NamedStyle(name="kpi_bad", font=Font(color="9C0006", bold=True),
        fill=PatternFill("solid", fgColor="FFC7CE"))

    s["kpi_warn"] = NamedStyle(name="kpi_warn", font=Font(color="9C6500", bold=True),
        fill=PatternFill("solid", fgColor="FFEB9C"))
    s["title"] = NamedStyle(name="title", font=Font(size=18, bold=True, color=TEAL_DARK))
    s["subtitle"] = NamedStyle(name="subtitle", font=Font(size=11, color=GREY))
    s["section_header"] = NamedStyle(name="section_header", font=Font(size=13, bold=True, color=TEAL_DARK),
        fill=PatternFill("solid", fgColor="E7F1EF"))
    s["section_header_center"] = NamedStyle(name="section_header_center", font=Font(size=13, bold=True, color=TEAL_DARK),
        fill=PatternFill("solid", fgColor="E7F1EF"), alignment=Alignment(horizontal="center"))

    # Register globally so we don't hit duplicate style errors
    seen = set()
    existing_styles = [x.name for x in wb._named_styles] if hasattr(wb, '_named_styles') else []
    for name, style in s.items():
        if style.name not in existing_styles:
            try:
                wb.add_style(style)
                existing_styles.append(style.name)
            except Exception:
                seen.add(name)
    return s

# ── Core Engine ──
class SmartWorkbookBuilder:
    def __init__(self, path, title=None, subtitle=None):
        self.path = path
        self.wb = Workbook()
        self.styles = _make_styles(self.wb)
        self._setup_sheet()
        self._write_cover(title or "", subtitle or "")

    def _setup_sheet(self):
        self.ws = self.wb.active
        self.ws.title = "Dashboard"
        self.ws.sheet_view.showGridLines = False
        self.ws.freeze_panes = "A3"
        self.ws.row_dimensions[1].height = 20
        self.ws.row_dimensions[2].height = 15

    def _write_cover(self, title, subtitle):
        self.ws["A1"] = title
        self.ws["A1"].style = self.styles["title"]
        self.ws.merge_cells("A1:F1")
        self.ws["A2"] = subtitle or ""
        self.ws["A2"].style = self.styles["subtitle"]
        self.ws.merge_cells("A2:F2")

    def add_section(self, header, col_count=4, start_row=None):
        row = start_row or self.ws.max_row + 2
        self.ws.cell(row=row, column=1, value=header)
        self.ws.cell(row=row, column=1).style = self.styles["section_header"]
        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
        return row

    def add_table(self, headers, data, start_row, table_name=None, has_total=False):
        col_count = len(headers)
        # Write headers
        for c, h in enumerate(headers, 1):
            cell = self.ws.cell(row=start_row, column=c, value=h)
            cell.style = self.styles["header"]

        # Write data
        for r_offset, row_data in enumerate(data, 1):
            for c, val in enumerate(row_data, 1):
                cell = self.ws.cell(row=start_row + r_offset, column=c, value=val)
                cell.style = self.styles["value"]

        # Auto column width
        for col in range(1, col_count + 1):
            max_len = max(
                len(str(self.ws.cell(row=r, column=col).value))
                for r in range(start_row, start_row + len(data))
            )
            self.ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)

        # Convert to Excel Table
        if table_name:
            table_ref = f"A{start_row}:{get_column_letter(col_count)}{start_row + len(data)}"
            tab = Table(displayName=table_name, ref=table_ref)
            style_info = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style_info
            self.ws.add_table(tab)

        return start_row + len(data) + 1

    def apply_kpi_formatting(self, column_letter, start_row=2, good_threshold=0.9, warn_threshold=0.75):
        max_row = self.ws.max_row
        data_range = f"{column_letter}{start_row}:{column_letter}{max_row}"

        # Green for above threshold
        self.ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator='greaterThanOrEqual', formula=[str(good_threshold)],
                       stopIfTrue=True, fill=PatternFill("solid", fgColor="C6EFCE"))
        )
        self.ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator='between', formula=[str(warn_threshold), str(good_threshold)],
                       stopIfTrue=True, fill=PatternFill("solid", fgColor="FFEB9C"))
        )
        self.ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator='lessThan', formula=[str(warn_threshold)],
                       stopIfTrue=True, fill=PatternFill("solid", fgColor="FFC7CE"))
        )

    def apply_currency_format(self, column_letter, start_row=2):
        max_row = self.ws.max_row
        for cell in self.ws[f"{column_letter}{start_row}:{column_letter}{max_row}"]:
            for c in cell:
                c.number_format = '#,##0.00'

    def apply_percentage_format(self, column_letter, start_row=2):
        max_row = self.ws.max_row
        for cell in self.ws[f"{column_letter}{start_row}:{column_letter}{max_row}"]:
            for c in cell:
                c.number_format = '0.00%'

    def apply_data_bars(self, column_letter, start_row=2, color="0F5C56"):
        max_row = self.ws.max_row
        rule = DataBarRule(start=None, end=None, color=color, showValue=True)
        self.ws.conditional_formatting.add(f"{column_letter}{start_row}:{column_letter}{max_row}", rule)

    def save(self):
        self.wb.save(self.path)
        return self.path


def generate_kpi_dashboard(data, output_path, title="OAKAI KPI Dashboard", subtitle=""):
    """
    Generate a fully formatted KPI dashboard workbook.

    Args:
        data: Dict mapping sheet names to dicts containing:
              {
                  "sections": [
                      {"header": str, "tables": [{"headers": list, "rows": list, "table_name": str?, "formats": dict?}]}
                  ]
              }
        output_path: Full path to save the .xlsx file
        title: Dashboard main title (shown on cover sheet)
        subtitle: Optional subtitle string

    Returns:
        dict with keys: 'path', 'sheets', 'timestamp'
    """
    wb_builder = SmartWorkbookBuilder(output_path, title, subtitle)
    sheets_created = []

    # Handle multi-sheet case
    if len(data) > 1:
        # First sheet already created ("Dashboard"), rename it
        wb_builder.ws.title = list(data.keys())[0]
        sheets_created.append(wb_builder.ws.title)

        for idx, (sheet_name, config) in enumerate(data.items()):
            if idx == 0:
                ws = wb_builder.ws  # Reuse first one
            else:
                ws = wb_builder.wb.create_sheet(sheet_name)
                sheets_created.append(sheet_name)

            current_row = 3  # Skip cover area
            for section in config.get("sections", []):
                section_row = wb_builder.add_section(section["header"], start_row=current_row)
                current_row = section_row + 1
                for table in section.get("tables", []):
                    current_row = wb_builder.add_table(
                        headers=table["headers"],
                        data=table["rows"],
                        start_row=current_row,
                        table_name=table.get("name")
                    )
                    # Apply formats
                    fmts = table.get("formats", {})
                    for col_idx, fmt in enumerate(fmts.values(), 1):
                        col_letter = get_column_letter(col_idx)
                        if fmt == "currency":
                            wb_builder.apply_currency_format(col_letter, start_row=current_row - len(table["rows"]))
                        elif fmt == "percentage":
                            wb_builder.apply_percentage_format(col_letter, start_row=current_row - len(table["rows"]))
                        elif fmt == "kpi":
                            wb_builder.apply_kpi_formatting(col_letter, start_row=current_row - len(table["rows"]))

    else:
        # Single sheet case
        sheet_name = list(data.keys())[0]
        wb_builder.ws.title = sheet_name
        sheets_created.append(sheet_name)
        current_row = 3
        for section in data[sheet_name].get("sections", []):
            section_row = wb_builder.add_section(section["header"], start_row=current_row)
            current_row = section_row + 1
            for table in section.get("tables", []):
                current_row = wb_builder.add_table(
                    headers=table["headers"],
                    data=table["rows"],
                    start_row=current_row,
                    table_name=table.get("name")
                )
                fmts = table.get("formats", {})
                for col_idx, fmt in enumerate(fmts.values(), 1):
                    col_letter = get_column_letter(col_idx)
                    if fmt == "currency":
                        wb_builder.apply_currency_format(col_letter, start_row=current_row - len(table["rows"]))
                    elif fmt == "percentage":
                        wb_builder.apply_percentage_format(col_letter, start_row=current_row - len(table["rows"]))
                    elif fmt == "kpi":
                        wb_builder.apply_kpi_formatting(col_letter, start_row=current_row - len(table["rows"]))

    # Save
    saved_path = wb_builder.save()
    ts = datetime.now(timezone.utc).isoformat()

    return {
        "path": saved_path,
        "sheets": sheets_created,
        "timestamp": ts
    }


# ── CLI Entry Point ──
if __name__ == "__main__":
    import argparse, json
    parser = argparse.ArgumentParser(description="Generate KPI dashboard from JSON spec")
    parser.add_argument("input_json", help="Path to JSON file describing dashboard structure")
    parser.add_argument("-o", "--output", default="/opt/data/workspace/OAKAI_KPI_Dashboard.xlsx")
    parser.add_argument("-t", "--title", default="OAKAI KPI Dashboard")
    parser.add_argument("-s", "--subtitle", default="")
    args = parser.parse_args()

    with open(args.input_json) as f:
        spec = json.load(f)

    result = generate_kpi_dashboard(
        data=spec,
        output_path=args.output,
        title=args.title,
        subtitle=args.subtitle
    )

    print(f"✓ Dashboard generated: {result['path']}")
    print(f"  Sheets: {', '.join(result['sheets'])}")
    print(f"  Timestamp: {result['timestamp']}")
    print(f"  Size: {os.path.getsize(result['path'])} bytes")
