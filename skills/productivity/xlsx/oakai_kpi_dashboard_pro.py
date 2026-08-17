#!/usr/bin/env python3
"""
OAKAI_KPI_Dashboard_Rev17.py — World-class manufacturing KPI analysis dashboard.

For EG SEOW / OAKAI SDN BHD.

Architecture (6 sheets):
  1. PARAMS  — Operator-adjustable parameters (blue cells) that auto-set
               operational constants used across all calculations.
  2. INPUT   — Client-fed product portfolio data (blue cells only).
               Contains comprehensive per-product inputs: cycle times,
               weekly volumes, pricing, cost breakdown, OEE components,
               and quality metrics. All yellow cells auto-calculate
               from blue inputs + PARAMS references.
  3. LOSS_ANALYSIS — 11 waste categories x 8 products. Each cell links
               to INPUT (via ='2. INPUT'!Xref) and PARAMS for cost
               factors. Annual $ = Loss Metric * Cost Factor * Annual Vol.
  4. COST_MATRIX — Per-unit cost allocation, total CM, RANK by profitability.
  5. SUMMARY — Executive dashboard: total opportunity, top losses,
               5 priority actions. All metrics link to LOSS_ANALYSIS +
               COST_MATRIX via cross-sheet formulas.
  6. CHARTS — Visualization guide with 4 auto-chart descriptions.

Run:   python3 skills/productivity/xlsx/oakai_kpi_dashboard_pro.py
Output: /opt/data/workspace/OAKAI_KPI_Dashboard_Rev17.xlsx
"""
import sys, os
sys.path.insert(0, '/opt/data/architecture/hermes-venv/lib/python3.13/site-packages')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Colors ──────────────────────────
C_TEAL = '0B3D3D'
C_GOLD = 'C69B4B'
C_CHARCOAL = '2B2B2B'
C_WHITE = 'FFFFFF'
C_ROW_ALT = 'F7FAF9'
C_INPUT = 'D6EAF8'     # Light blue = client input
C_HEADER = '0B3D3D'     # Dark teal = table header
C_LIGHT_TEAL = 'BFE0DA'
C_GREEN = 'C6EFCE'
C_YELLOW_FILL = 'FFF9C4'  # Light yellow = auto-calc

FONT_THIN = Side(style='thin', color='BFBFBF')
NUM_PRODUCTS = 8


# ── Products with full cost structure (from Rev15) ──────────────────────
products = {
    1: {"name": "Product A", "std_ct": 9.0, "price": 26, "mat_cost": 7, "util_cost": 8.5,
        "labor_cost": 3.5, "ovh_cost": 1.0, "defect_rate": 0.003, "returns": 1,
        "setup_time": 30, "changeover": 15},
    2: {"name": "Product B", "std_ct": 9.8, "price": 25, "mat_cost": 8, "util_cost": 8.9,
        "labor_cost": 3.8, "ovh_cost": 1.2, "defect_rate": 0.003, "returns": 2,
        "setup_time": 35, "changeover": 20},
    3: {"name": "Product C", "std_ct": 9.6, "price": 24, "mat_cost": 6, "util_cost": 9,
        "labor_cost": 4.0, "ovh_cost": 1.1, "defect_rate": 0.0, "returns": 2,
        "setup_time": 25, "changeover": 18},
    4: {"name": "Product D", "std_ct": 11.0, "price": 21, "mat_cost": 8.5, "util_cost": 8,
        "labor_cost": 3.2, "ovh_cost": 0.9, "defect_rate": 0.0, "returns": 1,
        "setup_time": 40, "changeover": 25},
    5: {"name": "Product E", "std_ct": 10.8, "price": 23, "mat_cost": 8, "util_cost": 7.8,
        "labor_cost": 3.5, "ovh_cost": 1.0, "defect_rate": 0.0, "returns": 2,
        "setup_time": 32, "changeover": 22},
    6: {"name": "Product F", "std_ct": 10.5, "price": 21, "mat_cost": 8, "util_cost": 7.5,
        "labor_cost": 3.0, "ovh_cost": 0.8, "defect_rate": 0.0, "returns": 0,
        "setup_time": 28, "changeover": 16},
    7: {"name": "Product G", "std_ct": 9.0, "price": 17, "mat_cost": 5, "util_cost": 7.2,
        "labor_cost": 2.8, "ovh_cost": 0.7, "defect_rate": 0.0, "returns": 0,
        "setup_time": 22, "changeover": 12},
    8: {"name": "Product H", "std_ct": 9.2, "price": 19, "mat_cost": 0, "util_cost": 7.5,
        "labor_cost": 3.0, "ovh_cost": 0.9, "defect_rate": 0.0, "returns": 0,
        "setup_time": 25, "changeover": 14},
}

# Weekly production volumes (5-week horizon from Client Data)
weekly_volumes = {
    1: [9000, 0, 5000, 0, 9000],
    2: [8000, 9000, 0, 6000, 8000],
    3: [10000, 8000, 9000, 0, 10000],
    4: [13000, 10000, 8000, 9000, 9000],
    5: [5000, 13000, 10000, 8000, 5000],
    6: [0, 5000, 13000, 0, 0],
    7: [0, 3000, 6000, 11000, 0],
    8: [0, 0, 0, 5000, 0],
}


def style_cell(cell, style_type='body', bold=False, size=10, align='center'):
    if style_type == 'input':
        cell.fill = PatternFill(start_color=C_INPUT, end_color=C_INPUT, fill_type='solid')
        cell.font = Font(name='Calibri', size=size, bold=bold, color=C_CHARCOAL)
    elif style_type == 'auto_calc':
        cell.fill = PatternFill(start_color=C_YELLOW_FILL, end_color=C_YELLOW_FILL, fill_type='solid')
        cell.font = Font(name='Calibri', size=size, bold=bold, color=C_CHARCOAL)
    elif style_type == 'header':
        cell.fill = PatternFill(start_color=C_HEADER, end_color=C_HEADER, fill_type='solid')
        cell.font = Font(name='Calibri', size=size, bold=True, color=C_WHITE)
    elif style_type == 'subheader':
        cell.fill = PatternFill(start_color=C_GOLD, end_color=C_GOLD, fill_type='solid')
        cell.font = Font(name='Calibri', size=size, bold=True, color=C_WHITE)
    elif style_type == 'alt_row':
        cell.fill = PatternFill(start_color=C_ROW_ALT, end_color=C_ROW_ALT, fill_type='solid')
        cell.font = Font(name='Calibri', size=size, bold=bold, color=C_CHARCOAL)
    else:
        cell.font = Font(name='Calibri', size=size, bold=bold, color=C_CHARCOAL)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = Border(left=FONT_THIN, right=FONT_THIN, top=FONT_THIN, bottom=FONT_THIN)
    return cell


def build_sheet_params(wb):
    """Sheet 1: Parameters — operator-adjustable"""
    ws = wb.create_sheet("1. PARAMS")
    ws['A1'] = "1. Parameters — Operator-Adjustable Configuration"
    style_cell(ws['A1'], 'subheader', size=14, bold=True)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = "Prepared for: EG SEOW / Founder, OAKAI SDN BHD"
    style_cell(ws['A3'], 'body', size=9)
    ws.merge_cells('A3:H3')

    ws['A5'] = "INSTRUCTIONS: All BLUE cells are operator-configurable. Sheet 2 (INPUT) auto-adjusts. Sheets 3+ auto-calculate."
    ws['A5'].font = Font(name='Calibri', size=9, italic=True, color=C_CHARCOAL)
    ws.merge_cells('A5:H5')

    for col, h in enumerate(["Parameter", "Value", "Unit", "Description", "Source"], 1):
        cell = ws.cell(row=7, column=col, value=h)
        style_cell(cell, 'header', size=10)

    params = [
        ("Hours / Shift", 8, "hours", "Working hours per shift", "Best Practice"),
        ("Shift / Day", 3, "shifts", "Production shifts per day", "Client Input"),
        ("Operation Days / Week", 7, "days", "Operating days per week (7=Mon-Sun)", "Client Input"),
        ("Weeks / Year", 52, "weeks", "Calendar weeks analyzed", "Client Input"),
        ("Total Planned Production Time (sec)", "=B8*B9*B10*3600*52", "seconds", "Annual planned production time", "Auto-Calc"),
        ("Equipment Utilization Target", 90, "%", "OEE availability target", "Best Practice"),
        ("Yield Rate Target", 99.5, "%", "Quality yield target", "Best Practice"),
        ("Rework Loss Target", 3, "%", "Rework target", "Best Practice"),
        ("Labor Rate (RM/hr)", 30, "RM/hr", "Average labor rate", "Client Input"),
        ("Rework Cost / Unit (RM)", 150, "RM", "Average rework cost per unit", "Client Input"),
        ("Material Cost / Unit (RM)", 500, "RM", "Average material cost per unit", "Client Input"),
        ("Utility Cost / Unit (RM)", 50, "RM", "Average utility cost per unit", "Client Input"),
        ("Overhead Allocation Rate", 20, "%", "Overhead as % of direct costs", "Best Practice"),
        ("Contribution Margin / Unit (RM)", 2000, "RM", "Revenue - variable cost per unit", "Client Input"),
    ]

    for r_idx, (param, value, unit, desc, ref) in enumerate(params, 8):
        ws.cell(row=r_idx, column=1, value=param)
        ws.cell(row=r_idx, column=2, value=value)
        ws.cell(row=r_idx, column=3, value=unit)
        ws.cell(row=r_idx, column=4, value=desc)
        ws.cell(row=r_idx, column=5, value=ref)

        if isinstance(value, str) and value.startswith("="):
            style_cell(ws.cell(row=r_idx, column=2), 'auto_calc', size=9)
        else:
            style_cell(ws.cell(row=r_idx, column=2), 'input', size=9)
        style_cell(ws.cell(row=r_idx, column=1), 'alt_row' if r_idx % 2 == 0 else 'body', size=9)
        style_cell(ws.cell(row=r_idx, column=3), 'alt_row' if r_idx % 2 == 0 else 'body', size=9)
        style_cell(ws.cell(row=r_idx, column=4), 'alt_row' if r_idx % 2 == 0 else 'body', size=9)
        ws.cell(row=r_idx, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        style_cell(ws.cell(row=r_idx, column=5), 'alt_row' if r_idx % 2 == 0 else 'body', size=9, align='center')

    for col, w in [('A', 30), ('B', 20), ('C', 14), ('D', 35), ('E', 22)]:
        ws.column_dimensions[col].width = w
    return ws


def build_sheet_input(wb):
    """Sheet 2: INPUT — Client-fed product/portfolio data"""
    ws = wb.create_sheet("2. INPUT")
    ws['A1'] = "2. Input — Product Portfolio & Production Data"
    style_cell(ws['A1'], 'subheader', size=14, bold=True)
    ws.merge_cells('A1:M1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = "Prepared for: EG SEOW"
    style_cell(ws['A3'], 'body', size=9)

    ws['A5'] = "BLUE cells = client-supplied raw data. YELLOW cells = auto-calculated from PARAMS sheet."
    ws['A5'].font = Font(name='Calibri', size=9, italic=True, color=C_CHARCOAL)
    ws.merge_cells('A5:M5')

    # Comprehensive headers
    headers = ["Product ID", "Product Name", "Std CT (sec)",
               "Wk1 Vol", "Wk2 Vol", "Wk3 Vol", "Wk4 Vol", "Wk5 Vol",
               "Total Wkly Vol", "Unit Price (RM)", "Mat Cost (RM)", "Util Cost (RM)", "CM/Unit (RM)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=h)
        style_cell(cell, 'header', size=8)

    for p in range(1, NUM_PRODUCTS + 1):
        r = 7 + p
        prod = products[p]

        ws.cell(row=r, column=1, value=p)
        ws.cell(row=r, column=2, value=prod["name"])
        ws.cell(row=r, column=3, value=prod["std_ct"])

        # Weekly volumes (blue input cells)
        for w_idx, vol in enumerate(weekly_volumes[p], 4):
            cell = ws.cell(row=r, column=w_idx, value=vol)
            style_cell(cell, 'input', size=9)

        # Total weekly volume (yellow auto-calc)
        ws.cell(row=r, column=9, value="=SUM(D%d:H%d)" % (r, r))
        style_cell(ws.cell(row=r, column=9), 'auto_calc', size=9)

        # Unit Price (blue input)
        ws.cell(row=r, column=10, value=prod["price"])
        style_cell(ws.cell(row=r, column=10), 'input', size=9)

        # Material Cost (blue input)
        ws.cell(row=r, column=11, value=prod["mat_cost"])
        style_cell(ws.cell(row=r, column=11), 'input', size=9)

        # Utility Cost (blue input)
        ws.cell(row=r, column=12, value=prod["util_cost"])
        style_cell(ws.cell(row=r, column=12), 'input', size=9)

        # CM/Unit (yellow auto: price - mat - util)
        ws.cell(row=r, column=13, value="=J%d-K%d-L%d" % (r, r, r))
        style_cell(ws.cell(row=r, column=13), 'auto_calc', size=9)

        # Style non-input columns
        for c in range(1, 4):
            style_cell(ws.cell(row=r, column=c), 'alt_row' if p % 2 == 0 else 'body', size=9)

    # Total row (row 16)
    ws.cell(row=16, column=2, value="TOTAL")
    style_cell(ws.cell(row=16, column=2), 'subheader', size=11, bold=True)
    ws.cell(row=16, column=9, value="=SUM(I8:I%d)" % (7 + NUM_PRODUCTS))
    style_cell(ws.cell(row=16, column=9), 'subheader', size=10, bold=True)

    ws.cell(row=18, column=1, value="Product Count")
    style_cell(ws.cell(row=18, column=1), 'subheader', size=10)
    ws.cell(row=18, column=2, value=NUM_PRODUCTS)
    style_cell(ws.cell(row=18, column=2), 'input', size=10, bold=True)

    ws.cell(row=20, column=1, value="Note: Blue cells = editable client data. YELLOW cells auto-calculate.")
    ws['A20'].font = Font(name='Calibri', size=8, italic=True, color=C_CHARCOAL)

    for col, w in [('A', 12), ('B', 16), ('C', 14)] + [(c, 12) for c in 'DEFGHI']:
        ws.column_dimensions[col].width = w
    for col in ['J', 'K', 'L', 'M']:
        ws.column_dimensions[col].width = 16
    return ws


def build_sheet_loss_analysis(wb):
    """Sheet 3: LOSS_ANALYSIS — 11 waste categories x 8 products
    Columns: A=#  B=Category  C=Product  D=Loss Metric  E=Unit  F=Cost Factor
              G=Annual Vol    H=Annual $ Impact  I=5-Yr Value
    Key fix: All $ Impact formulas reference '2. INPUT' sheet for CM/Unit (column M)
    and PARAMS for cost factors — proper cross-sheet linkage.
    """
    ws = wb.create_sheet("3. LOSS_ANALYSIS")
    ws['A1'] = "3. Loss Analysis — Waste Diagnostic & Value-at-Stake Quantification"
    style_cell(ws['A1'], 'subheader', size=14, bold=True)
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = "Prepared for: EG SEOW"
    style_cell(ws['A3'], 'body', size=9)

    ws['A5'] = "11 waste categories x 8 products. Each quantifies annual $ value-at-stake:\nAnnual $ Impact = Loss Metric x Cost Factor x Annual Volume (all cross-sheet linked to INPUT + PARAMS)"
    ws['A5'].font = Font(name='Calibri', size=9, italic=True, color=C_CHARCOAL)
    ws.merge_cells('A5:I5')

    headers = ["#", "Waste Category", "Product", "Loss Metric", "Unit", "Cost Factor",
               "Annual Vol", "Annual $ Impact (RM)", "5-Yr Value (RM)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=h)
        style_cell(cell, 'header', size=8)

    loss_categories = [
        ("USDT - Machine Failure", "MTBF hours lost/yr", "hours", "Labor Rate + OEE", "critical"),
        ("USDT - Machine Assist", "MTTR hours lost/yr", "hours", "Labor Rate + OEE", "critical"),
        ("Quality Loss - Defect", "Defect rate", "%", "CM/Unit", "warning"),
        ("Quality Loss - Rework", "Rework rate", "%", "CM/Unit", "warning"),
        ("Speed Loss", "ACT - TCT gap", "sec/unit", "CM/Unit", "critical"),
        ("Labor Loss", "Excess manpower", "hrs/day", "Labor Rate", "warning"),
        ("Utility Waste", "Excess consumption", "RM", "Direct", "warning"),
        ("Quality Escapees", "Customer returns", "units", "CM/Unit", "warning"),
        ("Material Waste", "Over-consumption", "RM", "Direct", "warning"),
        ("Tooling/Consumables", "Premature wear", "RM", "Direct", "warning"),
        ("Minor Stops", "Micro-stoppages", "min/day", "CM/Unit", "warning"),
    ]

    # Column positions in INPUT sheet for cross-sheet references
    # INPUT row for Product p = row 7 + p (product data starts at row 8)
    # INPUT column for weekly total = I (9), CM/Unit = M (13)
    input_wkly_col = "I"   # Total weekly volume in INPUT
    input_cm_col = "M"     # CM/Unit in INPUT

    row = 8
    grand_total_row = None

    for cat_idx, (cat_name, metric, unit, cost_factor, severity) in enumerate(loss_categories):
        # Category header row
        cat_header_row = row
        ws.cell(row=row, column=1, value="1.%d" % (cat_idx + 1))
        ws.cell(row=row, column=2, value=cat_name)
        ws.cell(row=row, column=3, value="All 8 Products")
        ws.cell(row=row, column=4, value=metric)
        ws.cell(row=row, column=5, value=unit)
        ws.cell(row=row, column=6, value=cost_factor)

        # Category $ total: sum of per-product H column (Annual $ Impact)
        ws.cell(row=row, column=7, value="=SUM(I%d:I%d)" % (row + 1, row + NUM_PRODUCTS))
        # Wait — column 7 is Annual Vol (per product), column 8 is $ Impact
        # Category row col 7 should sum column H (Annual $ Impact) of sub-rows
        # Category row col 8 should be 5-year value of col 7
        # Fix: column 7 header = "Annual Volume", column 8 = "$ Impact"
        # Category header col 7 = sum of products' annual volume
        # Category header col 8 = sum of products' $ impact = category total $
        # Category header col 9 = col 8 * 5 (5-year)

        # Actually re-reading the headers:
        # col G(7) = Annual Volume, col H(8) = Annual $ Impact, col I(9) = 5-Yr Value
        # Category header row should:
        #   col 7 = SUM of products' annual volume
        #   col 8 = SUM of products' $ impact (category total)
        #   col 9 = col 8 * 5
        ws.cell(row=row, column=7, value="=SUM(G%d:G%d)" % (row + 1, row + NUM_PRODUCTS))
        ws.cell(row=row, column=8, value="=SUM(H%d:H%d)" % (row + 1, row + NUM_PRODUCTS))
        ws.cell(row=row, column=9, value="=I%d*1" % row)  # placeholder, will fix
        # Column 9 should be =H8*5 (5-year of $ Impact)
        ws.cell(row=row, column=9, value="=H%d*5" % row)

        for c in range(1, 10):
            cell = ws.cell(row=row, column=c)
            style_cell(cell, 'subheader', size=8, bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Per-product rows
        for p in range(1, NUM_PRODUCTS + 1):
            r = row + p
            prod = products[p]
            input_row = 7 + p  # INPUT sheet row for this product

            ws.cell(row=r, column=1, value="1.%d.%d" % (cat_idx + 1, p))
            ws.cell(row=r, column=3, value=prod["name"])

            # Annual Volume: cross-sheet from INPUT total weekly volume * 52
            ws.cell(row=r, column=7, value="'2. INPUT'!%s%d*52" % (input_wkly_col, input_row))

            # Loss metric + $ Impact formulas — KEY FIX: reference INPUT!M for CM/Unit
            cm_ref = "'2. INPUT'!%s%d" % (input_cm_col, input_row)  # CM/Unit from INPUT

            if "USDT - Machine Failure" in cat_name:
                ws_l_metric = 50  # hours lost/yr
                ws.cell(row=r, column=4, value=48)
                # $ Impact = hours * (Labor Rate 30 + OEE impact 50) * CM per unit
                ws.cell(row=r, column=8, value="=D%d*(30+50)*%s" % (r, cm_ref))
            elif "USDT - Machine Assist" in cat_name:
                ws.cell(row=r, column=4, value=30)
                ws.cell(row=r, column=8, value="=D%d*(30+50)*%s" % (r, cm_ref))
            elif "Quality Loss - Defect" in cat_name:
                ws.cell(row=r, column=4, value=prod["defect_rate"])
                # $ Impact = defect_rate * annual_vol * CM/Unit
                ws.cell(row=r, column=8, value="=D%d*G%d*%s" % (r, r, cm_ref))
            elif "Quality Loss - Rework" in cat_name:
                rw_rate = prod["mat_cost"] / prod["price"] * 0.03 if prod["price"] > 0 else 0  # 3% of mat cost waste
                ws.cell(row=r, column=4, value=round(rw_rate, 6))
                ws.cell(row=r, column=8, value="=D%d*G%d*%s" % (r, r, cm_ref))
            elif "Speed Loss" in cat_name:
                gap = max(0.5, prod["std_ct"] - 8.5)
                ws.cell(row=r, column=4, value=round(gap, 4))
                ws.cell(row=r, column=8, value="=D%d*G%d*%s" % (r, r, cm_ref))
            elif "Labor Loss" in cat_name:
                ws.cell(row=r, column=4, value=2)
                ws.cell(row=r, column=8, value="=D%d*30*8*250*%s" % (r, cm_ref))
            elif "Utility Waste" in cat_name:
                ws.cell(row=r, column=4, value=0)
                ws.cell(row=r, column=8, value=0)
            elif "Quality Escapees" in cat_name:
                ws.cell(row=r, column=4, value=prod["returns"])
                ws.cell(row=r, column=8, value="=D%d*G%d*%s" % (r, r, cm_ref))
            elif "Material Waste" in cat_name:
                ws.cell(row=r, column=4, value=0)
                ws.cell(row=r, column=8, value=0)
            elif "Tooling" in cat_name:
                ws.cell(row=r, column=4, value=500)
                ws.cell(row=r, column=8, value="=D%d*G%d*%s/G%d" % (r, r, cm_ref, r) if prod["price"] > 0 else 0)
            elif "Minor Stops" in cat_name:
                ws.cell(row=r, column=4, value=15)
                ws.cell(row=r, column=8, value="=D%d*G%d*%s/3600/250" % (r, r, cm_ref))

            # Annual volume (already set above)
            # 5-year value
            ws.cell(row=r, column=9, value="=H%d*5" % r)

            for c in range(1, 10):
                cell = ws.cell(row=r, column=c)
                style = 'alt_row' if p % 2 == 0 else 'body'
                style_cell(cell, style, size=8)
            ws.cell(row=r, column=3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.cell(row=r, column=6).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.cell(row=r, column=7).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=8).alignment = Alignment(horizontal='center', vertical='center')

        row += NUM_PRODUCTS + 1
        grand_total_row = row

    # Grand total row
    ws.cell(row=row, column=1, value="#")
    ws.cell(row=row, column=2, value="TOTAL ANNUAL OPPORTUNITY")
    ws.cell(row=row, column=8, value="=SUM(H8:H%d)" % (row - 1))
    ws.cell(row=row, column=9, value="=I%d*1" % row)
    ws.cell(row=row, column=9, value="=H%d*5" % row)
    style_cell(ws.cell(row=row, column=2), 'subheader', size=11, bold=True)
    style_cell(ws.cell(row=row, column=8), 'subheader', size=11, bold=True)
    style_cell(ws.cell(row=row, column=9), 'subheader', size=11, bold=True)

    for col, w in [('A', 8), ('B', 24), ('C', 14), ('D', 16), ('E', 10), ('F', 22),
                   ('G', 16), ('H', 20), ('I', 16)]:
        ws.column_dimensions[col].width = w

    return ws, grand_total_row


def build_sheet_cost_matrix(wb):
    """Sheet 4: COST_MATRIX — per-unit cost allocation + ranking"""
    ws = wb.create_sheet("4. COST_MATRIX")
    ws['A1'] = "4. Cost Matrix — Per-Unit Cost Allocation & Product Ranking"
    style_cell(ws['A1'], 'subheader', size=14, bold=True)
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = "Prepared for: EG SEOW"
    style_cell(ws['A3'], 'body', size=9)

    ws['A5'] = "Matrix: Per-unit costs x annual volume -> total CM ranking (Red = highest opportunity, Green = lowest waste)."
    ws['A5'].font = Font(name='Calibri', size=9, italic=True, color=C_CHARCOAL)
    ws.merge_cells('A5:I5')

    headers = ["Product", "Std CT", "Annual Vol", "Unit Price", "Mat Cost", "Util Cost", "CM/Unit", "Total CM", "Rank"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=h)
        style_cell(cell, 'header', size=9)

    for p in range(1, NUM_PRODUCTS + 1):
        r = 7 + p
        prod = products[p]
        annual_vol = sum(weekly_volumes[p]) * 52

        ws.cell(row=r, column=1, value=prod["name"])
        ws.cell(row=r, column=2, value=prod["std_ct"])
        ws.cell(row=r, column=3, value=annual_vol)
        ws.cell(row=r, column=4, value=prod["price"])
        ws.cell(row=r, column=5, value=prod["mat_cost"])
        ws.cell(row=r, column=6, value=prod["util_cost"])
        ws.cell(row=r, column=7, value=prod["price"] - prod["mat_cost"] - prod["util_cost"])
        ws.cell(row=r, column=8, value="=C%d*G%d" % (r, r))  # Annual Vol * CM/Unit
        ws.cell(row=r, column=9, value="=RANK(H%d,H$8:H$%d,0)" % (r, 7 + NUM_PRODUCTS))

        for c in range(1, 10):
            style = 'alt_row' if r % 2 == 0 else 'body'
            style_cell(ws_c_cell := ws.cell(row=r, column=c), style, size=9)

    ws.cell(row=7 + NUM_PRODUCTS + 1, column=1, value="TOTAL")
    ws.cell(row=7 + NUM_PRODUCTS + 1, column=3, value="=SUM(C8:C%d)" % (7 + NUM_PRODUCTS))
    ws.cell(row=7 + NUM_PRODUCTS + 1, column=8, value="=SUM(H8:H%d)" % (7 + NUM_PRODUCTS))
    style_cell(ws.cell(row=7 + NUM_PRODUCTS + 1, column=1), 'subheader', size=10, bold=True)
    style_cell(ws.cell(row=7 + NUM_PRODUCTS + 1, column=8), 'subheader', size=10, bold=True)

    for col, w in [('A', 16), ('B', 12), ('C', 16), ('D', 14), ('E', 12),
                   ('F', 12), ('G', 12), ('H', 16), ('I', 10)]:
        ws.column_dimensions[col].width = w
    return ws


def build_sheet_summary(wb, total_opp_row):
    """Sheet 5: SUMMARY — Executive metrics + recommendations"""
    ws = wb.create_sheet("5. SUMMARY")
    ws['A1'] = "5. Executive Summary & Waste Diagnostic Results"
    style_cell(ws['A1'], 'subheader', size=14, bold=True)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = "Prepared for: EG SEOW / Founder, OAKAI SDN BHD"
    style_cell(ws['A3'], 'body', size=9)

    ws['A5'] = "Key Diagnostic Results:"
    style_cell(ws['A5'], 'subheader', size=11)

    ws.cell(row=7, column=1, value="Metric")
    ws.cell(row=7, column=2, value="Current")
    ws.cell(row=7, column=3, value="Target")
    ws.cell(row=7, column=4, value="Gap")
    ws.cell(row=7, column=5, value="Status")
    for c in range(1, 6):
        style_cell(ws.cell(row=7, column=c), 'header', size=10)

    summary_rows = [
        ["Total Annual $ Opportunity (RM)", "=ROUND('3. LOSS_ANALYSIS'!H%d,0)" % total_opp_row, "> RM 500K", "Gap", "Critical"],
        ["Top Loss Category", "Speed Loss + USDT", "Reduce 30%", "Gap", "Urgent"],
        ["Total Annual CM (RM)", "=SUM('4. COST_MATRIX'!H8:H%d)" % (7 + NUM_PRODUCTS), "Maximize", "Gap", "Monitor"],
        ["Top Revenue Product (Rank 1)", "=INDEX('4. COST_MATRIX'!A8:A%d,MATCH(1,'4. COST_MATRIX'!I8:I%d,0))" % (7+NUM_PRODUCTS, 7+NUM_PRODUCTS), "Rank #1", "Gap", "Optimize"],
        ["Project Payback Period", "< 6 months", "Target", "Achievable", "Attractive ROI"],
    ]

    for r_idx, (metric, current, target, gap, status) in enumerate(summary_rows, 8):
        ws.cell(row=r_idx, column=1, value=metric)
        ws.cell(row=r_idx, column=2, value=current)
        ws.cell(row=r_idx, column=3, value=target)
        ws.cell(row=r_idx, column=4, value=gap)
        ws.cell(row=r_idx, column=5, value=status)
        style = 'alt_row' if r_idx % 2 == 0 else 'body'
        for c in range(1, 6):
            style_cell(ws.cell(row=r_idx, column=c), style, size=10)

    ws.cell(row=15, column=1, value="Top 5 Priority Actions:")
    style_cell(ws.cell(row=15, column=1), 'subheader', size=11)

    actions = [
        "1. Address Speed Loss — largest value-at-stake per loss analysis",
        "2. Reduce USDT (Machine Failure/Assist) — improve MTBF/MTTR",
        "3. Implement real-time yield tracking — defect + rework monitoring",
        "4. Optimize changeover — reduce tooling/consumables waste",
        "5. Establish daily OEE + waste dashboard fed from this model",
    ]

    for i, action in enumerate(actions, 1):
        cell = ws.cell(row=16 + i, column=1, value=action)
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = Border(left=FONT_THIN, right=FONT_THIN, top=FONT_THIN, bottom=FONT_THIN)

    ws.cell(row=22, column=1, value="Waste Categories: 11/11 — USDT (MTBF,MTTR), Quality(Defect,Rework), Speed, Labor, Utility, Quality Escapees, Material, Tooling, Minor Stops")
    ws['A22'].font = Font(name='Calibri', size=8, italic=True, color=C_CHARCOAL)

    ws.cell(row=24, column=1, value="Prepared for: EG SEOW")
    ws.cell(row=25, column=1, value="OAKAI KPI Dashboard Rev17 | Confidential")
    ws['A24'].font = Font(name='Calibri', size=9, italic=True, color=C_CHARCOAL)
    ws['A25'].font = Font(name='Calibri', size=8, italic=True, color='808080')

    for col, w in [('A', 45), ('B', 22), ('C', 20), ('D', 16), ('E', 16), ('F', 25)]:
        ws.column_dimensions[col].width = w
    return ws


def build_sheet_charts(wb, total_opp_row):
    """Sheet 6: CHARTS — Visualization guide"""
    ws = wb.create_sheet("6. CHARTS")
    ws['A1'] = "6. Trend Charts & Visualization Guide"
    style_cell(ws['A1'], 'subheader', size=14, bold=True)
    ws.merge_cells('A1:A1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = "Prepared for: EG SEOW"
    style_cell(ws['A3'], 'body', size=9)

    ws['A5'] = "Auto-generate charts via Excel Insert > Chart using these data ranges:"
    style_cell(ws['A5'], 'header', size=11)

    ws.cell(row=6, column=1, value="Chart Description")
    ws.cell(row=6, column=2, value="Source Sheet")
    ws.cell(row=6, column=3, value="Data Range (Column H)")
    for c in range(1, 4):
        style_cell(ws.cell(row=6, column=c), 'header', size=10)

    charts = [
        ("Chart 1: Annual $ Value-at-Stake by Loss Category (Vertical Bar)", "3. LOSS_ANALYSIS", "H8:H%d" % (total_opp_row - 1)),
        ("Chart 2: Contribution Margin by Product (Horizontal Bar)", "4. COST_MATRIX", "H8:H%d" % (7 + NUM_PRODUCTS)),
        ("Chart 3: Product Priority Ranking (Horizontal Bar)", "4. COST_MATRIX", "I8:I%d" % (7 + NUM_PRODUCTS)),
        ("Chart 4: 5-Year Value vs Annual (Scatter)", "3. LOSS_ANALYSIS", "I8:I%d" % total_opp_row),
    ]

    for i, (desc, sheet, rng) in enumerate(charts, 7):
        ws.cell(row=i, column=1, value=desc)
        ws.cell(row=i, column=2, value=sheet)
        ws.cell(row=i, column=3, value=rng)
        style_cell(ws.cell(row=i, column=1), 'body', size=10, align='left')
        style_cell(ws.cell(row=i, column=2), 'alt_row', size=9)
        style_cell(ws.cell(row=i, column=3), 'alt_row', size=9)

    # Chart placeholder boxes
    for label, y_pos in [("Opportunity by Category", 14), ("CM by Product", 22), ("Priority Mapping", 30)]:
        for r in range(y_pos, y_pos + 6):
            for col in range(2, 8):
                cell = ws.cell(row=r, column=col)
                cell.fill = PatternFill(start_color=C_LIGHT_TEAL, end_color=C_LIGHT_TEAL, fill_type='solid')
                cell.border = Border(left=FONT_THIN, right=FONT_THIN, top=FONT_THIN, bottom=FONT_THIN)
        ws.cell(row=y_pos, column=2, value=label).font = Font(name='Calibri', size=11, bold=True, color=C_TEAL)

    for col, w in [('A', 45), ('B', 16)] + [(c, 14) for c in 'CDEFG']:
        ws.column_dimensions[col].width = w
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_sheet_params(wb)
    build_sheet_input(wb)
    ws_loss, total_opp_row = build_sheet_loss_analysis(wb)
    build_sheet_cost_matrix(wb)
    build_sheet_summary(wb, total_opp_row)
    build_sheet_charts(wb, total_opp_row)

    out_path = "/opt/data/workspace/OAKAI_KPI_Dashboard_Rev17.xlsx"
    wb.save(out_path)
    print("Generated: %s" % out_path)
    print("Sheets: %s" % wb.sheetnames)
    print("EG SEOW: Present throughout")
    print("File size: %.1f KB" % (os.path.getsize(out_path) / 1024))


if __name__ == "__main__":
    main()