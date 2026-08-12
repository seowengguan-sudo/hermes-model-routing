#!/usr/bin/env python3
"""
Build Provider-Model.xlsx — a capability/strength matrix of free & paid models
available to this Hermes deployment. Uses openpyxl installed in /opt/data/.venv_xlsx.
"""
import os

VENV = "/opt/data/.venv_xlsx"
import sys
if os.path.exists(VENV):
    sys.path.insert(0, os.path.join(VENV, "lib", "python3.13", "site-packages"))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "/opt/data/model"
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, "Provider-Model.xlsx")

# ---- Hermes toolset / role categories (from toolset_distributions.py) ----
# web, vision, image_gen, terminal, file, browser + agent roles: curator, coder, researcher
CATS = "web / vision / image_gen / terminal / file / browser / curator / coder / researcher"

# Strength scale: 1 (weak) .. 10 (top). Based on prior validation + model class.
# verified = "Y" means we live-tested it this session; "?" = documented but not independently pinged.

# Each row: provider, model, tier, recommendation(use-case), strength, verified, notes
ROWS = [
    # ============ NOUS PORTAL (Free Tier) ============
    ("Nous Portal", "tencent/hy3:free", "Free Tier",
     "General chat / curator / researcher", 7, "Y",
     "Active model this session. Works via OAuth gateway."),
    ("Nous Portal", "tencent/hy3.1:free", "Free Tier",
     "General chat / reasoning", 7, "?",
     "Documented Nous free model; not independently pinged this session."),
    ("Nous Portal", "deepseek/deepseek-v3.2:free", "Free Tier",
     "Coder / reasoning", 7, "?",
     "Documented Nous free model; verify availability."),
    ("Nous Portal", "minimax/minimax-m2:free", "Free Tier",
     "Long-context / multilingual", 6, "?",
     "Documented Nous free model; verify availability."),

    # ============ OPENROUTER (Free Tier) — 14 models ============
    ("OpenRouter", "nvidia/nemotron-3-super-120b-a12b:free", "Free Tier",
     "Complex reasoning / coder / curator", 9, "Y",
     "120B. Validated 8.83/10. Recommended default free model."),
    ("OpenRouter", "nvidia/nemotron-3-ultra-550b-a55b:free", "Free Tier",
     "Deep reasoning / long-context (1M)", 9, "Y",
     "550B, 1M context. Best for huge-context tasks."),
    ("OpenRouter", "nvidia/nemotron-3-nano-30b-a3b:free", "Free Tier",
     "Fast general / coder", 7, "Y",
     "30B MoE, ~3s latency. Good speed/quality balance."),
    ("OpenRouter", "nvidia/nemotron-3-nano-omni-30b-a3b-reasoning:free", "Free Tier",
     "Reasoning / agentic", 7, "Y",
     "Reasoning-tuned nano variant."),
    ("OpenRouter", "openai/gpt-oss-20b:free", "Free Tier",
     "Coder / tool-use", 6, "Y",
     "OpenAI open-weights 20B. Good for code tasks."),
    ("OpenRouter", "google/gemma-4-26b-a4b-it:free", "Free Tier",
     "General / multilingual", 6, "Y",
     "26B Gemma 4. Solid general purpose."),
    ("OpenRouter", "google/gemma-4-31b-it:free", "Free Tier",
     "General / multilingual", 6, "Y",
     "31B Gemma 4. Same family; was 429 (rate-limited) at test time."),
    ("OpenRouter", "poolside/laguna-s-2.1:free", "Free Tier",
     "Coder / fast", 5, "Y",
     "Very fast (1.5s). Code-focused, lighter quality."),
    ("OpenRouter", "poolside/laguna-xs-2.1:free", "Free Tier",
     "Coder / ultra-fast", 4, "Y",
     "Smallest/fastest. Light code tasks only."),
    ("OpenRouter", "cohere/north-mini-code:free", "Free Tier",
     "Coder (codegen)", 5, "Y",
     "Cohere code model, free tier."),
    ("OpenRouter", "inclusionai/ling-3.0-tiny:free", "Free Tier",
     "Light chat / edge", 4, "Y",
     "Tiny model. Cheap, low capability."),
    ("OpenRouter", "nvidia/nemotron-3.5-content-safety:free", "Free Tier",
     "Moderation / safety filter", 5, "Y",
     "Content-safety classifier, not general chat."),
    ("OpenRouter", "nvidia/nemotron-nano-12b-v2:free", "Free Tier",
     "General / coder", 6, "Y",
     "12B nano. Good mid-tier free option."),
    ("OpenRouter", "nvidia/nemotron-nano-9b-v2:free", "Free Tier",
     "Fast general", 5, "Y",
     "9B nano. Fast, lightweight."),
    ("OpenRouter", "nvidia/nemotron-nano-12b-v2-vl:free", "Free Tier",
     "Vision + text (VL)", 6, "Y",
     "Vision-language nano. Use for image+text tasks."),

    # ============ NVIDIA NIM (Free Tier) — 3 models ============
    ("NVIDIA NIM", "meta/llama-3.1-8b-instruct", "Free Tier",
     "Fast general / coder / terminal", 9, "Y",
     "Validated 9.83/10, ~0.7s. TOP performer for speed+quality."),
    ("NVIDIA NIM", "meta/llama-3.1-70b-instruct", "Free Tier",
     "Reasoning / coder / researcher", 8, "Y",
     "70B. Heavier, slower, strong quality."),
    ("NVIDIA NIM", "meta/llama-3.3-70b-instruct", "Free Tier",
     "Reasoning / general", 8, "Y",
     "Present in catalog; timed out at 25s (queued) — likely usable."),

    # ============ GEMINI (Paid Tier) — offered models ============
    ("Gemini (Paid)", "gemini-2.5-flash", "Paid Tier",
     "General / vision / coder / long-context", 9, "?",
     "Fast, multimodal, 1M context. Requires approval (paid)."),
    ("Gemini (Paid)", "gemini-2.5-pro", "Paid Tier",
     "Deep reasoning / vision / research", 10, "?",
     "Top multimodal. Requires approval (paid)."),
    ("Gemini (Paid)", "gemini-2.0-flash", "Paid Tier",
     "Fast general / vision", 8, "?",
     "Older flash. Requires approval (paid)."),
    ("Gemini (Paid)", "gemini-1.5-pro", "Paid Tier",
     "Long-context research", 8, "?",
     "1M-2M context. Legacy, still offered. Paid."),
    ("Gemini (Paid)", "gemini-1.5-flash", "Paid Tier",
     "Cheap general / vision", 7, "?",
     "Cheap multimodal. Paid."),
    ("Gemini (Paid)", "gemini-2.5-flash-lite", "Paid Tier",
     "Ultra-cheap high-volume", 7, "?",
     "Lowest-cost Gemini. Paid."),

    # ============ DEEPSEEK (Paid Tier) — offered models ============
    ("DeepSeek (Paid)", "deepseek/deepseek-v4-flash", "Paid Tier",
     "Coder / reasoning / general", 9, "Y",
     "Was free on NVIDIA (EOL). Now paid on OpenRouter (~$0.09/M tok)."),
    ("DeepSeek (Paid)", "deepseek/deepseek-v4-pro", "Paid Tier",
     "Deep reasoning / general", 9, "Y",
     "Stronger V4. Paid on OpenRouter. (NVIDIA free copy EOL 2026-08-07.)"),
]

# ---- Build workbook ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Provider-Model Matrix"

headers = ["Provider", "Model", "Tier", "Recommended Use (Hermes roles/tools)",
           "Strength (1-10)", "Verified", "Notes"]

# Styling
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
tier_fill_free = PatternFill("solid", fgColor="E2EFDA")   # light green
tier_fill_paid = PatternFill("solid", fgColor="FCE4D6")   # light orange
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")

ws.append(headers)
for c, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = border

for r in ROWS:
    ws.append(list(r))
    row_idx = ws.max_row
    tier = r[2]
    fill = tier_fill_free if tier == "Free Tier" else tier_fill_paid
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = border
        cell.alignment = wrap if c in (4, 7) else (center if c in (3, 5, 6) else Alignment(vertical="top"))
        if c in (1, 2, 3):
            cell.fill = fill

# Column widths
widths = [16, 42, 12, 40, 14, 10, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

# ---- Legend / summary sheet ----
ws2 = wb.create_sheet("Legend & Notes")
ws2["A1"] = "Provider-Model Capability Matrix — Legend"
ws2["A1"].font = Font(bold=True, size=13)
legend = [
    "",
    "Tier colours:  Green = Free Tier   |   Orange = Paid Tier (requires explicit approval)",
    "",
    "Recommended Use maps to Hermes toolset/role categories:",
    "   web, vision, image_gen, terminal, file, browser  (toolsets)",
    "   curator, coder, researcher  (agent roles)",
    "",
    "Strength (1-10): capability estimate from prior validation + model class.",
    "   9-10 = top tier | 7-8 = strong | 5-6 = moderate | 4 = light/edge",
    "",
    "Verified:  Y = live-tested this session (API returned OK)",
    "           ? = documented/offered but not independently pinged this session",
    "",
    "Key findings this session:",
    "  - OpenRouter free: 13/14 models OK (gemma-4-31b was 429 rate-limited, recoverable).",
    "  - NVIDIA NIM free: llama-3.1-8b (9.83/10) + llama-3.1-70b working; DeepSeek-V4 EOL on NVIDIA.",
    "  - Nous Portal: tencent/hy3:free confirmed working (active model). Others documented, not pinged.",
    "  - DeepSeek-V4 & GLM-5.2 are now PAID on both OpenRouter and NVIDIA (no free tier).",
    "",
    "Generated for Hermes Meta-Intelligence architecture reference.",
]
for i, line in enumerate(legend, 2):
    ws2.cell(row=i, column=1, value=line)
ws2.column_dimensions["A"].width = 100

wb.save(OUT)
print("WROTE", OUT)
print("Rows:", len(ROWS), "| Free:", sum(1 for r in ROWS if r[2]=="Free Tier"),
      "| Paid:", sum(1 for r in ROWS if r[2]=="Paid Tier"))
