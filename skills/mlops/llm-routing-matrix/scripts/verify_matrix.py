#!/usr/bin/env python3
"""verify_matrix.py — sanity-check a Provider-Model routing workbook.

Usage: python3 verify_matrix.py <path-to-xlsx>

Checks every category's 4-slot sequence in the Category_Sequence sheet:
  - capability mismatch (vision/approval/mcp must hold the right model type)
  - duplicate adjacent slots (slot1 == slot2)
  - paid-default violation (heavy model used as default paid slot)
Exits non-zero if any issue is found.
"""
import sys
import openpyxl

# Model -> capability role (extend as needed)
ROLE = {
    'nvidia/nemotron-nano-12b-v2-vl': 'vision',
    'nvidia/nemotron-3.5-content-safety': 'moderation',
    'openai/gpt-oss-20b': 'tool-use',
    'deepseek/deepseek-v4-pro': 'heavy',
}
# category -> required role of slot that is NOT "—"
REQUIRED = {'vision': 'vision', 'approval': 'moderation', 'mcp': 'tool-use'}
HEAVY_DEFAULT_PAID = {'deepseek/deepseek-v4-pro', 'gemini-2.5-pro'}


def main(path):
    wb = openpyxl.load_workbook(path)
    assert 'Category_Sequence' in wb.sheetnames, 'Category_Sequence sheet missing'
    ws = wb['Category_Sequence']
    issues = []
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        if not cat:
            continue
        slots = [ws.cell(r, c).value for c in (2, 3, 4, 5)]
        # duplicate adjacent
        if slots[0] and slots[0] == slots[1]:
            issues.append(f'{cat}: slot1 == slot2 ({slots[0]})')
        # capability mismatch
        if cat in REQUIRED:
            role = REQUIRED[cat]
            # find first non-"—" slot with a model
            placed = None
            for s in slots:
                if s and '—' not in str(s) and s in ROLE:
                    placed = s
                    break
            if placed and ROLE.get(placed) != role:
                issues.append(f'{cat}: slot holds {placed} but needs {role}')
        # paid default must not be heavy
        paid = slots[3] or ''
        if any(h in paid for h in HEAVY_DEFAULT_PAID):
            issues.append(f'{cat}: heavy model used as DEFAULT paid ({paid}) — reserve for super-complex')
    if issues:
        print('ISSUES FOUND:')
        for i in issues:
            print(' -', i)
        sys.exit(1)
    print('All sanity checks passed.')


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('usage: verify_matrix.py <workbook.xlsx>')
        sys.exit(2)
    main(sys.argv[1])
