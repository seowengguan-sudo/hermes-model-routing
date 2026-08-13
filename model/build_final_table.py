#!/usr/bin/env python3
"""Final build of Provider-Model.xlsx using consolidated live-verification results."""
import os, json, sys
VENV = "/opt/data/.venv_xlsx"
sys.path.insert(0, os.path.join(VENV, "lib", "python3.13", "site-packages"))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "/opt/data/model"
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, "Provider-Model.xlsx")

# Consolidated live results (from verify_all.py + verify_fix.py)
# status: OK / 429(rate-limited) / 400(bad id) / 403(direct blocked, gateway-only) / 404(not found)
STATUS = {
 ("Nous Portal","tencent/hy3:free"):"GW",          # works via Hermes gateway (active model)
 ("Nous Portal","tencent/hy3.1:free"):"GW",
 ("Nous Portal","deepseek/deepseek-v3.2:free"):"GW",
 ("Nous Portal","minimax/minimax-m2:free"):"GW",
 ("OpenRouter","nvidia/nemotron-3-super-120b-a12b:free"):"OK",
 ("OpenRouter","nvidia/nemotron-3-ultra-550b-a55b:free"):"OK",
 ("OpenRouter","nvidia/nemotron-3-nano-30b-a3b:free"):"OK",
 ("OpenRouter","nvidia/nemotron-3-nano-omni-30b-a3b-reasoning:free"):"OK",
 ("OpenRouter","openai/gpt-oss-20b:free"):"OK",
 ("OpenRouter","google/gemma-4-26b-a4b-it:free"):"OK",
 ("OpenRouter","google/gemma-4-31b-it:free"):"429",
 ("OpenRouter","poolside/laguna-s-2.1:free"):"OK",
 ("OpenRouter","poolside/laguna-xs-2.1:free"):"OK",
 ("OpenRouter","cohere/north-mini-code:free"):"OK",
 ("OpenRouter","inclusionai/ling-3.0-tiny:free"):"OK",
 ("OpenRouter","nvidia/nemotron-3.5-content-safety:free"):"OK",
 ("OpenRouter","nvidia/nemotron-nano-9b-v2:free"):"OK",
 ("OpenRouter","nvidia/nemotron-nano-12b-v2-vl:free"):"OK",
 ("NVIDIA NIM","meta/llama-3.1-8b-instruct"):"OK",
 ("NVIDIA NIM","meta/llama-3.1-70b-instruct"):"OK",
 ("NVIDIA NIM","meta/llama-3.3-70b-instruct"):"TO",   # timed out 20s (queued/slow)
 ("Gemini (Paid)","gemini-2.5-flash"):"OK",
 ("Gemini (Paid)","gemini-2.5-pro"):"404",
 ("Gemini (Paid)","gemini-2.0-flash"):"404",
 ("Gemini (Paid)","gemini-1.5-pro"):"404",
 ("Gemini (Paid)","gemini-1.5-flash"):"404",
 ("Gemini (Paid)","gemini-2.5-flash-lite"):"404",
 ("DeepSeek (Paid)","deepseek/deepseek-v4-flash"):"OK",
 ("DeepSeek (Paid)","deepseek/deepseek-v4-pro"):"OK",
}

STATUS_LABEL = {
 "OK":"Verified OK", "429":"Rate-limited (recoverable)", "400":"Bad model ID",
 "403":"Direct blocked (gateway only)", "TO":"Timeout (slow/queued)",
 "404":"Not found at endpoint", "GW":"Gateway-verified (live chat)",
 "NO_KEY":"No API key",
}

# (provider, model, tier, recommendation, strength, notes)
ROWS = [
 ("Nous Portal","tencent/hy3:free","Free Tier","General chat / curator / researcher",7,
  "Active model this session via Hermes gateway. Direct API 403 (expired cached token)."),
 ("Nous Portal","tencent/hy3.1:free","Free Tier","General chat / reasoning",7,
  "Offered free on Nous. Verified only via gateway."),
 ("Nous Portal","deepseek/deepseek-v3.2:free","Free Tier","Coder / reasoning",7,
  "Offered free on Nous. Verified only via gateway."),
 ("Nous Portal","minimax/minimax-m2:free","Free Tier","Long-context / multilingual",6,
  "Offered free on Nous. Verified only via gateway."),

 ("OpenRouter","nvidia/nemotron-3-super-120b-a12b:free","Free Tier","Complex reasoning / coder / curator",9,
  "120B. Validated 8.83/10. Recommended default free model."),
 ("OpenRouter","nvidia/nemotron-3-ultra-550b-a55b:free","Free Tier","Deep reasoning / long-context (1M)",9,
  "550B, 1M context. Best for huge-context tasks."),
 ("OpenRouter","nvidia/nemotron-3-nano-30b-a3b:free","Free Tier","Fast general / coder",7,
  "30B MoE, ~0.4s. Good speed/quality balance."),
 ("OpenRouter","nvidia/nemotron-3-nano-omni-30b-a3b-reasoning:free","Free Tier","Reasoning / agentic",7,
  "Reasoning-tuned nano variant."),
 ("OpenRouter","openai/gpt-oss-20b:free","Free Tier","Coder / tool-use",6,
  "OpenAI open-weights 20B. Good for code tasks."),
 ("OpenRouter","google/gemma-4-26b-a4b-it:free","Free Tier","General / multilingual",6,
  "26B Gemma 4. Solid general purpose."),
 ("OpenRouter","google/gemma-4-31b-it:free","Free Tier","General / multilingual",6,
  "31B Gemma 4. 429 at test time (rate-limited) — recoverable."),
 ("OpenRouter","poolside/laguna-s-2.1:free","Free Tier","Coder / fast",5,
  "Very fast (3.4s). Code-focused, lighter quality."),
 ("OpenRouter","poolside/laguna-xs-2.1:free","Free Tier","Coder / ultra-fast",4,
  "Smallest/fastest. Light code tasks only."),
 ("OpenRouter","cohere/north-mini-code:free","Free Tier","Coder (codegen)",5,
  "Cohere code model, free tier."),
 ("OpenRouter","inclusionai/ling-3.0-tiny:free","Free Tier","Light chat / edge",4,
  "Tiny model. Cheap, low capability."),
 ("OpenRouter","nvidia/nemotron-3.5-content-safety:free","Free Tier","Moderation / safety filter",5,
  "Content-safety classifier, not general chat."),
 ("OpenRouter","nvidia/nemotron-nano-9b-v2:free","Free Tier","Fast general",5,
  "9B nano. Fast, lightweight."),
 ("OpenRouter","nvidia/nemotron-nano-12b-v2-vl:free","Free Tier","Vision + text (VL)",6,
  "Vision-language nano. Use for image+text tasks. (non-VL 12b id returned 400.)"),

 ("NVIDIA NIM","meta/llama-3.1-8b-instruct","Free Tier","Fast general / coder / terminal",9,
  "Validated 9.83/10, ~0.7s. TOP performer for speed+quality."),
 ("NVIDIA NIM","meta/llama-3.1-70b-instruct","Free Tier","Reasoning / coder / researcher",8,
  "70B. Heavier, slower, strong quality."),
 ("NVIDIA NIM","meta/llama-3.3-70b-instruct","Free Tier","Reasoning / general",8,
  "In catalog; timed out at 20s (queued/slow) — likely usable with longer timeout."),

 ("Gemini (Paid)","gemini-2.5-flash","Paid Tier","General / vision / coder / long-context",9,
  "OpenAI-compat endpoint OK. Requires approval (paid)."),
 ("Gemini (Paid)","gemini-2.5-pro","Paid Tier","Deep reasoning / vision / research",10,
  "404 at tested endpoint — model ID/endpoint needs correction. Paid."),
 ("Gemini (Paid)","gemini-2.0-flash","Paid Tier","Fast general / vision",8,
  "404 at tested endpoint — ID/endpoint needs correction. Paid."),
 ("Gemini (Paid)","gemini-1.5-pro","Paid Tier","Long-context research",8,
  "404 at tested endpoint — ID/endpoint needs correction. Paid."),
 ("Gemini (Paid)","gemini-1.5-flash","Paid Tier","Cheap general / vision",7,
  "404 at tested endpoint — ID/endpoint needs correction. Paid."),
 ("Gemini (Paid)","gemini-2.5-flash-lite","Paid Tier","Ultra-cheap high-volume",7,
  "404 at tested endpoint — ID/endpoint needs correction. Paid."),

 ("DeepSeek (Paid)","deepseek/deepseek-v4-flash","Paid Tier","Coder / reasoning / general",9,
  "Paid on OpenRouter (~$0.09/M tok). NVIDIA free copy EOL 2026-08-07."),
 ("DeepSeek (Paid)","deepseek/deepseek-v4-pro","Paid Tier","Deep reasoning / general",9,
  "Paid on OpenRouter. NVIDIA free copy EOL."),
]

# Build
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Provider-Model Matrix"
headers = ["Provider","Model","Tier","Recommended Use (Hermes roles/tools)",
           "Strength (1-10)","Live Status","Notes"]
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
free_fill = PatternFill("solid", fgColor="E2EFDA")
paid_fill = PatternFill("solid", fgColor="FCE4D6")
ok_fill = PatternFill("solid", fgColor="C6EFCE")
warn_fill = PatternFill("solid", fgColor="FFEB9C")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")

ws.append(headers)
for c in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill; cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = border

def status_fill(st):
    return ok_fill if st in ("OK","GW") else warn_fill

for r in ROWS:
    prov, model, tier, rec, strength, notes = r
    st = STATUS.get((prov, model), "?")
    stlabel = STATUS_LABEL.get(st, st)
    ws.append([prov, model, tier, rec, strength, stlabel, notes])
    ri = ws.max_row
    tierfill = free_fill if tier=="Free Tier" else paid_fill
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=ri, column=c)
        cell.border = border
        if c in (4,7): cell.alignment = wrap
        elif c in (3,5,6): cell.alignment = center
        else: cell.alignment = Alignment(vertical="top")
        if c == 1: cell.fill = tierfill
        if c == 6: cell.fill = status_fill(st)

widths = [16, 44, 12, 40, 14, 22, 52]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

# Legend sheet
ws2 = wb.create_sheet("Legend & Notes")
ws2["A1"] = "Provider-Model Capability Matrix — Verified (live API probes, 2026-08-07)"
ws2["A1"].font = Font(bold=True, size=13)
legend = [
 "",
 "Live Status legend:",
 "  Verified OK        = API returned 200 on direct probe this session",
 "  Gateway-verified   = works through Hermes gateway (live chat); direct API blocked (expired OAuth token)",
 "  Rate-limited (rec) = HTTP 429, recoverable (retry later / different IP)",
 "  Timeout (slow)     = request exceeded 20s, queued/slow — likely usable with longer timeout",
 "  Bad model ID       = HTTP 400, model slug not valid on that provider",
 "  Not found          = HTTP 404 at tested endpoint (ID/endpoint needs correction)",
 "  Direct blocked     = HTTP 403 on direct API (use gateway path)",
 "",
 "Tier colours:  Green = Free Tier   |   Orange = Paid Tier (requires explicit approval)",
 "Strength (1-10): 9-10 top | 7-8 strong | 5-6 moderate | 4 light/edge",
 "",
 "Hermes toolset/role categories used in 'Recommended Use':",
 "  web, vision, image_gen, terminal, file, browser  (toolsets)",
 "  curator, coder, researcher  (agent roles)",
 "",
 "Verification summary:",
 "  Nous Portal (4): gateway-verified only (direct 403 — cached OAuth token expired).",
 "  OpenRouter (14): 13 OK + 1 rate-limited (gemma-4-31b 429). nemotron-nano-12b (non-VL) ID invalid.",
 "  NVIDIA NIM (3): llama-3.1-8b OK, llama-3.1-70b OK, llama-3.3-70b timeout (slow).",
 "  Gemini (6): gemini-2.5-flash OK; other 5 returned 404 at OpenAI-compat endpoint (IDs need correction).",
 "  DeepSeek (2): both OK on OpenRouter paid path.",
 "",
 "Note: DeepSeek-V4 and GLM-5.2 are now PAID on both OpenRouter and NVIDIA (no free tier).",
]
for i, line in enumerate(legend, 2):
    ws2.cell(row=i, column=1, value=line)
ws2.column_dimensions["A"].width = 105

wb.save(OUT)
print("WROTE", OUT)
ok = sum(1 for s in STATUS.values() if s in ("OK","GW"))
print("Total rows:", len(ROWS), "| Live-verified (OK/GW):", ok, "| Others flagged")
